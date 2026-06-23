"""
Admin Export Service — FR-AD-02 (export)
==========================================
Generates CSV and styled multi-sheet XLSX reports.
Requires: openpyxl>=3.1.0
"""
import io
import csv

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, GradientFill
from openpyxl.utils import get_column_letter

# ── Style Constants ────────────────────────────────────────────────────────────
_HEADER_FILL   = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
_ALT_ROW_FILL  = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
_HEADER_FONT   = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_BODY_FONT     = Font(name="Calibri", size=10)
_THIN_BORDER   = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
_CENTER = Alignment(horizontal="center", vertical="center")
_RIGHT  = Alignment(horizontal="right",  vertical="center")
_LEFT   = Alignment(horizontal="left",   vertical="center")


def _style_header_row(ws, col_count: int) -> None:
    for c in range(1, col_count + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill   = _HEADER_FILL
        cell.font   = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER


def _style_data_rows(ws, col_count: int, numeric_cols: list = None) -> None:
    numeric_cols = numeric_cols or []
    for row_idx in range(2, ws.max_row + 1):
        fill = _ALT_ROW_FILL if row_idx % 2 == 0 else None
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.font   = _BODY_FONT
            cell.border = _THIN_BORDER
            if fill:
                cell.fill = fill
            cell.alignment = _RIGHT if c in numeric_cols else _LEFT


def _autofit(ws, min_width: int = 10, max_width: int = 40) -> None:
    for col in ws.columns:
        width = max(
            len(str(cell.value or "")) for cell in col
        )
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(width + 3, min_width), max_width)


# ── CSV Export ─────────────────────────────────────────────────────────────────

def export_kpis_csv(cursor) -> bytes:
    """Flat CSV of platform KPIs."""
    from services.admin_kpi_service import get_platform_kpis
    data = get_platform_kpis(cursor)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Metric", "Value"])
    writer.writerow(["Platform Health Score", f"{data['health_score']}/100"])
    writer.writerow(["Health Label", data["health_label"]])
    for key, val in data["kpis"].items():
        writer.writerow([key.replace("_", " ").title(), val])
    return output.getvalue().encode("utf-8")


# ── XLSX Export (multi-sheet) ──────────────────────────────────────────────────

def export_kpis_xlsx(cursor) -> bytes:
    """
    Multi-sheet Excel workbook:
      Sheet 1 — Platform KPIs
      Sheet 2 — Sellers Report
      Sheet 3 — Category Analytics
    """
    wb = Workbook()

    # ── Sheet 1: Platform KPIs ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Platform KPIs"
    ws1.row_dimensions[1].height = 22

    from services.admin_kpi_service import get_platform_kpis
    kpi_data = get_platform_kpis(cursor)

    ws1.append(["Metric", "Value", "Note"])
    ws1.append(["Platform Health Score", kpi_data["health_score"], f"{kpi_data['health_label']} (out of 100)"])
    ws1.append(["", "", ""])  # spacer
    for key, val in kpi_data["kpis"].items():
        ws1.append([key.replace("_", " ").title(), val, ""])

    _style_header_row(ws1, 3)
    _style_data_rows(ws1, 3, numeric_cols=[2])
    _autofit(ws1)

    # ── Sheet 2: Sellers Report ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Sellers Report")
    headers2 = [
        "Business Name", "Category", "Email", "Status",
        "Active", "Churn Score", "Risk Level",
        "Rating", "Revenue (EGP)", "B2B Orders",
        "Fulfillment Hrs", "Profile %", "Fraud Score"
    ]
    ws2.append(headers2)

    from services.admin_seller_service import get_sellers_report
    for s in get_sellers_report(cursor, sort_by="revenue"):
        ws2.append([
            s["business_name"],
            s["category"],
            s["email"],
            s["approval_status"],
            "Yes" if s["is_active"] else "No",
            s["churn_risk_score"],
            s["risk_level"],
            s["avg_customer_rating"],
            s["total_revenue"],
            s["b2b_orders_total"],
            s["avg_fulfillment_hours"],
            s["profile_completeness"],
            s["avg_fraud_score"],
        ])

    _style_header_row(ws2, len(headers2))
    _style_data_rows(ws2, len(headers2), numeric_cols=[6, 8, 9, 10, 11, 12, 13])
    _autofit(ws2)

    # ── Sheet 3: Category Analytics ───────────────────────────────────────────
    ws3 = wb.create_sheet("Category Analytics")
    headers3 = [
        "Category", "Products", "Approved", "Pending",
        "Approval %", "Avg Price", "Total Purchases",
        "Total Revenue", "Avg Rating", "Quality Score",
        "Low Stock", "Sellers"
    ]
    ws3.append(headers3)

    from services.admin_category_service import get_category_analytics
    for c in get_category_analytics(cursor):
        ws3.append([
            c["category_name"],
            c["total_products"],
            c["approved_products"],
            c["pending_products"],
            c["approval_rate_pct"],
            c["avg_price"],
            c["total_purchases"],
            c["total_revenue"],
            c["avg_rating"],
            c["avg_quality_score"],
            c["low_stock_count"],
            c["seller_count"],
        ])

    _style_header_row(ws3, len(headers3))
    _style_data_rows(ws3, len(headers3), numeric_cols=list(range(2, 13)))
    _autofit(ws3)

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()
